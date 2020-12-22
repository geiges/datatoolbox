#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
This module contains all relevant tools regarding the use of excel


@author: ageiges
"""
import os
import re
import string
import pandas as pd
import tqdm
import platform

from openpyxl import load_workbook


#%% Defintions
OS = platform.system() #'win32' , linux, #Darwin


REQUIRED_SETUP_FIELDS = ['filePath',
                         'fileName',
                         'sheetName',
                         'timeIdxList',
                         'spaceIdxList']

SP_ARG = 3
TI_ARG = 2
DT_ARG = 4

REG_EXCEL_RANGE = re.compile('^[A-Z]{1,3}[0-9]{1,3}:[A-Z]{1,3}[0-9]{1,3}$')
REG_EXCEL_ROW   = re.compile('^[0-9]{1,2}$')
REG_EXCEL_COL   = re.compile('^[A-Z]{1,2}$')

REG_FIND_ROWS   = re.compile('^[A-Z]{1,3}([0-9]{1,3}):[A-Z]{1,3}([0-9]{1,3})$')
REG_FIND_COLS   = re.compile('^([A-Z]{1,3})[0-9]{1,3}:([A-Z]{1,3})[0-9]{1,3}$')





    

#%% Private functions
def _isColRange(string):
#    if REG_EXCEL_COL.match(string):
#        return True
    match = REG_FIND_COLS.search(string)
    if match:
        if match.group(1) == match.group(2):
            return True
    
    return False
        
def _isRowRange(string):
#    if REG_EXCEL_ROW.match(string):
#        return True
    match = REG_FIND_ROWS.search(string)
    if match:
        if match.group(1) == match.group(2):
            return True
    
    return False

def _isRow(string):
    if REG_EXCEL_ROW.match(string):
        return True
    else:
        return False
    
def _isCol(string):
    if REG_EXCEL_COL.match(string):
        return True
    else:
        return False
#%%
def _iterTime(timeIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):
    if _isRowRange(timeIdxString):
        # colum setup
        for timeCell in wksheet[timeIdxString][0]:
            xlsCol = timeCell.col_idx
            timeIdx = timeCell.value
            #print(xlsCol, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    
    elif _isRow(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
            xlsCol = timeCell.col_idx
            timeIdx = timeCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    elif _isCol(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
            xlsRow = timeCell.row
            timeIdx = timeCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
            
    elif _isColRange(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
#            print(timeCell[0].row)
            xlsRow = timeCell[0].row
            timeIdx = timeCell[0].value
            #print(xlsRow, timeIdx)
#            print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed
        timeIdx = timeIdxString
        yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

def _iterSpace(spaceIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):
    if _isRowRange(spaceIdxString):
        # colum setup
        for spaceCell in wksheet[spaceIdxString][0]:
            xlsCol = spaceCell.col_idx
            spaceIdx = spaceCell.value
            #print(xlsCol, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isRow(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsCol = spaceCell.col_idx
            spaceIdx = spaceCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    
    elif _isCol(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsRow = spaceCell.row
            spaceIdx = spaceCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
            
    elif _isColRange(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsRow = spaceCell[0].row
            spaceIdx = spaceCell[0].value
            #print(xlsRow, spaceIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed string
        print('fixed')
        spaceIdx = spaceIdxString
        yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
        
def _iterData(dataIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):

    #print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
    if _isRowRange(dataIdxString):
        # colum setup
        for dataCell in wksheet[dataIdxString][0]:
            xlsCol = dataCell.col_idx
            dataIdx = dataCell.value
            #print(xlsCol, timeIdx, dataIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isRow(dataIdxString):
        for dataCell in wksheet[dataIdxString]:
            xlsCol = dataCell.col_idx
            dataIdx = dataCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    
    
    elif _isCol(dataIdxString):
#        print('_isCol')
#        print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
    
        for i, dataCell in enumerate(wksheet[dataIdxString]):
#            print(dataCell)
#            print(dataCell.col_idx)
#            xlsRow = dataCell.row
            xlsCol = dataCell.col_idx
            if i == 0:
                dataIdx = dataCell.value
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
            
    elif _isColRange(dataIdxString):
#        print('_isColRange')
        for dataCell in wksheet[dataIdxString]:
            #print(dataCell)
            xlsRow = dataCell[0].row
            if dataCell[0].value is not None:
                dataIdx = dataCell[0].value.replace('\ufeff','')
            else:
                dataIdx = None
            #print(xlsRow, timeIdx)
            yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed
        dataIdx = dataIdxString
        yield  [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

def _identifyCountry(string):
    """
    Used various methods to find best fitting country alpha3 iso code
    """
    
    import datatoolbox as dt
    #numeric ISO code
    try:
        numISO = float(string)
        mask = numISO == dt.mapp.countries.codes['numISO']
        if mask.any():
            return dt.mapp.countries.index[mask][0]
    except:
        pass
    if len(string) == 2:
        mask = string == dt.mapp.countries.codes['alpha2']
        if mask.any():
            return dt.mapp.countries.codes.index[mask][0]
    
    if len(string) == 3:
         
        if string.upper() in dt.mapp.countries.codes.index:
            return string.upper()
    
    try: 
        coISO = dt.getCountryISO(string)
        return coISO
    except:
        print('not matching country found')    
        return None

def _res2Excel(resFull, countryList):
    import datatoolbox as dt
    tableSet = dt.TableSet()
    
    yearRange = (pd.np.inf, -pd.np.inf)
    
    ID_list = list()
    for ID in resFull:
        table = dt.getTable(ID)
        if 'Gg' in table.meta['unit']:
            table = table.convert(table.meta['unit'].replace('Gg','Mt'))
        tableSet.add(table)
        ID_list.append(table.ID)
        
        minYears = min(yearRange[0], table.columns.min())
        maxYears = max(yearRange[1], table.columns.max())
        yearRange = (minYears, maxYears)
        
    for country in countryList:
        
        coISO = _identifyCountry(country)
        
        if coISO is None:
            coISO = country
        
        outDf = pd.DataFrame(columns = ['Source', 'Entity', 'Category', 'Scenario', 'Unit'] + list(range(yearRange[0], yearRange[1]+1)))
        
        i =0
        for ID in ID_list:
            if coISO in tableSet[ID].index:
                years =  tableSet[ID].columns[~tableSet[ID].loc[coISO].isna()]
                outDf.loc[i,years] = tableSet[ID].loc[coISO,years]
                outDf.loc[i,['Source', 'Entity','Category', 'Scenario', 'Unit']] = [tableSet[ID].meta[x] for x in ['source', 'entity','category', 'scenario', 'unit']]
                i = i+1
        
        outDf = outDf.loc[:,~outDf.isnull().all(axis=0)]
        outDf.to_excel('extract_'+  coISO + '.xlsx')
 
def _str2float(x):
    if isinstance(x,float) or isinstance(x,int) or x is None:
        return x
    if '%' in x:
        return float(x.replace('%',''))*100
    else:
        if x.startswith('#') or x == '':
            return np.nan
#        try:
        print(x)
        return float(x)          
#%% Functions
        
def alphaCol2Num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) +1 
    return num-1

def colNameToNum(name):
    pow = 1
    colNum = 0
    for letter in name[::-1]:
            colNum += (int(letter, 36) -9) * pow
            pow *= 26
    return colNum -1

def compare_excel_files(file1, file2, eps=1e-6):
    """
    This function compares two excel files and creates a file visualising the differences
    
    """
    
    def report_diff(x):
        try:
#            print(x)
            x = x.astype(float)
#            print(x)
            return x[1] if (abs(x[0] - x[1]) < eps) or  pd.np.any(pd.isnull(x))  else '{0:.2f} -> {1:.2f}'.format(*x)
        except:
            return x[1] if (x[0] == x[1]) or  pd.np.any(pd.isnull(x))  else '{} -> {}'.format(*x)
            #print(x)
            
    xlFile1 = pd.ExcelFile(file1)
    sheetNameList1 = set(xlFile1.sheet_names)
    
    xlFile2 = pd.ExcelFile(file2)
    sheetNameList2 = set(xlFile2.sheet_names)
    
    writer = pd.ExcelWriter("diff_temp.xlsx")
    
    for sheetName in sheetNameList2.intersection(sheetNameList1):
        data1  = pd.read_excel(xlFile1, sheet_name=sheetName)
        data2  = pd.read_excel(xlFile2, sheet_name=sheetName)
        data1 = data1.replace(pd.np.nan, '').astype(str)
        data2 = data2.replace(pd.np.nan, '').astype(str)
        data1 = data1.apply(lambda x: x.str.strip())
        data2 = data2.apply(lambda x: x.str.strip())
        diff_panel = pd.Panel(dict(df1=data1,df2=data2))
        diff_output = diff_panel.apply(report_diff, axis=0)

        diff_output.to_excel(writer,sheet_name=sheetName)
    
        workbook  = writer.book
        worksheet = writer.sheets[sheetName]
        highlight_fmt = workbook.add_format({'font_color': '#FF0000', 'bg_color':'#B1B3B3'})
    
        grey_fmt = workbook.add_format({'font_color': '#8d8f91'})
        
        worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
                                                'criteria': 'containing',
                                                'value':'->',
                                                'format': highlight_fmt})
        worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
                                                'criteria': 'not containing',
                                                'value':'->',
                                                'format': grey_fmt})
        
    writer.close()
    os.system('libreoffice ' + "diff_temp.xlsx")
    os.system('rm diff_temp.xlsx')


def excel_to_pandas_idx(index):
    [alpha, num, _] = re.split(r'(\d+)',index)    
    return (int(num)-1, alphaCol2Num(alpha))

def getAllSheetNames(filePath):
    xlFile = pd.ExcelFile(filePath)
    sheetNameList = xlFile.sheet_names
    xlFile.close()
    return sheetNameList 


def getCountryExtract(countryList, sourceList='all'):
    """
    Extracts all available data in the database to an excel file
    """
    import datatoolbox as dt
    
    if sourceList == 'all':
        sourceList = list(dt.core.DB.sources.index)
    
    if not isinstance(countryList,list):
        countryList = [countryList]
    if not isinstance(sourceList,list):
        sourceList = [sourceList]
            
    #%%
    resFull = list()
    sourceList.sort()
    for source in tqdm.tqdm(sourceList):
        print(source)
        newList = list(dt.find(source=source).index)    
        newList.sort()
        resFull = resFull + newList

    _res2Excel(resFull, countryList)


def pandasStr2floatPercent(X):
    return([_str2float(x) for x in X])  
    
    
#%% Classes
    
class Excel_Reader():
    """
    Reader clase that uses an setup information provided in the excel file
    to extract data
    """
    
    def __init__(self, 
                 setup=None, 
                 fileName=None, 
                 overwrite=False, 
                 setupSheet = 'OUTPUT'):
        
        self.overwrite = overwrite
        self.setupSheet= setupSheet
        if setup:
            self.setup = setup
            fileSetup = True
        else:
            self.setup = dict()
            self.setup['fileName'] = fileName
            fileSetup = False
    
    def close(self):
        self.wb.close()
        
    def openSourceFile(self):
        if OS == 'Linux':
            os.system('libreoffice ' + self.setup['fileName'])
        elif OS == 'Darwin':
            os.system('open -a "Microsoft Excel" ' + self.setup['fileName'])        
   
    def getSetups(self):
        self.wb_read = load_workbook(self.setup['fileName'], data_only=True)
        self.wb = load_workbook(self.setup['fileName'], data_only=True)
        setup = dict()
        setup['fileName'] = self.setup['fileName']
        
        
        mapping = pd.read_excel(self.setup['fileName'], sheet_name =self.setupSheet)
        
        
        
        setupDict = {'sheetName'     : 'sheetName',
                     'timeIdxList'   : 'timeCells',
                     'spaceIdxList'  : 'spaceCells',
                     'dataID'        : 'dataCells',
                     'unit'          : 'unit',
                     'unitTo'        : 'unitTo'}

        setupDict = {'sheetName'     : 'Sheetname',
                     'timeIdxList'   : 'Time',
                     'spaceIdxList'  : 'Region',
                     'dataID'        : 'Variable',
                     'unit'          : 'Unit',
                     'unitTo'        : 'UnitTo',
                     'scenario'      : 'Scenario',
                     'source'        : 'Source'}
        
        for i,setupMapp in mapping.iterrows():
            for key in setupDict.keys():
                setup[key] = str(setupMapp[setupDict[key]])
                if setup[key] == 'nan':
                    setup[key] = None
                    
            yield(dict(setup))
            
#        setupSheet = self.wb[self.setupSheet]
#        for row in setupSheet.rows:
#            if row[0].value == 'sheet':
#                continue
#            
#            setup['sheetName']   = row[0].value
#            setup['timeIdxList']  = row[1].value
#            setup['spaceIdxList'] = row[2].value
#            setup['dataID']      = row[3].value
#            setup['unit']        = row[4].value
#            
#            for key in setup.keys():
#                if isinstance(setup[key],str):
#                    setup[key] = setup[key].replace('\ufeff','')
#            yield(setup)
        
    def read_data(self):
        replaceDict = {'EU': 'EU28'}
        from shutil import copyfile
        from copy import copy
        import mapping as mapp
        import datatoolbox as dt

        wb = load_workbook(self.setup['fileName'], data_only=True)
        
        validSpatialIDs = mapp.getValidSpatialIDs()
        self.setupList = list()
        
        tablesToReturn = dt.TableSet()
        
        for setup in self.getSetups():
            if isinstance(setup['timeIdxList'], int):
                setup['timeIdxList'] = str(setup['timeIdxList'])
            if isinstance(setup['spaceIdxList'], int):
                setup['spaceIdxList'] = str(setup['spaceIdxList'])
            print(setup)
            self.setupList.append(copy(setup))
            args =  None, None, None, None, None
            wksSheet = wb[setup['sheetName']]

            
            
            args =  [None, None, None, None, None]
            for argsTime in _iterTime(setup['timeIdxList'], wksSheet, *args):
                
                if argsTime[TI_ARG] is None:
                    print('No time defintion found')
                    print(setup['timeIdxList'])

                    continue
                print(argsTime[TI_ARG])
                print(setup['spaceIdxList'])
                for argsSpace in _iterSpace(setup['spaceIdxList'], wksSheet, *argsTime):
                    if argsSpace[SP_ARG] is None:
                        print('Not space defintion found')
                        continue
                    
                    if argsSpace[SP_ARG] not in validSpatialIDs:
                        print('No valid ISO code...')
                        
                        
                        if argsSpace[SP_ARG] in replaceDict.keys():
                            newID = replaceDict[argsSpace[SP_ARG]]
                        else:
                            newID = util.getCountryISO(argsSpace[SP_ARG])
                        

                        if newID in validSpatialIDs:
                            argsSpace[SP_ARG] = newID
                            print('Set iso code to: ' + argsSpace[SP_ARG])
                        else:
                            print('No ISO code found')
                            print(argsSpace[SP_ARG] + ' not found')
                            continue
                        
                    print(argsSpace)
                    for argData in _iterData(setup['dataID'], wksSheet, *argsSpace):
                        if argData[DT_ARG] is None:
                            print('No fitting dataID code found')
                            continue

                        meta = {'entity': argData[DT_ARG],
                                                      'unit' : setup['unit'],
                                                      'category':'',
                                                      'scenario' : setup['scenario'],
                                                      'source' : setup['source']}
#                        print(meta)
                        ID = core._createDatabaseID(meta)
                        
                        if ID not in tablesToReturn.keys():
                            table = Datatable()
                            table.meta = {'entity': argData[DT_ARG],
                                                      'unit' : setup['unit'],
                                                      'category':'',
                                                      'scenario' : setup['scenario'],
                                                      'source' : setup['source']}

                            tablesToReturn.add(table)
#                        print(argData)
                        value = self._readValue(wksSheet, xlsRow=argData[0], xlsCol=argData[1])

                        if isinstance(value, pd.DataFrame):
                            value = pandasStr2floatPercent(value)
                        else:
                            if value != '#VALUE!':
                                value = _str2float(value)
#                        print(value)
                        
#                        if  setup['unit'] == '%':
#                            value = value*100
#                        sdf
                        tablesToReturn[ID].loc[argData[SP_ARG],argData[TI_ARG]] = value
#                            print('success')
#                        iCount +=1
                            
#                        except Exception as e:
##                            print(e)
#                            pass
#                    self.wb.save(self.setup['fileName'])
#                            import pdb
#                            pdb.set_trace()
                        
        for table in tablesToReturn:
            table.columns = table.columns.astype(int)
            if '%' in table.meta['unit']:
                 table.loc[:,:] = table.loc[:,:]*100
                 tablesToReturn[table.ID] = table
                        
        return tablesToReturn
#            wb.save(self.setup['fileName'])
#            print('{} items inserted'.format(iCount))
#            
#            wb.close()

    def _readValue(self, wrkSheet, xlsRow, xlsCol):
#        if self.overwrite or pd.isna():
#            wrkSheet.cell(row=xlsRow, column=xlsCol, value = value)
        return wrkSheet.cell(row=xlsRow, column=xlsCol).value
#    def _writeMultipleIndicators(self, setup):
#        ws_readValues = load_workbook(self.setup['fileName'], data_only=True)[setup['sheetName']]
#        try:
#            tableIDs = [cell.value for cell in ws_readValues[setup['dataID']]]
#        except:
#            tableIDs = [cell[0].value for cell in ws_readValues[setup['dataID']]]
#            
#            tables = dt.getTables(tableIDs)
#            
#            REG_FIND_ROWS.search(setup['timeIdxList'])
        
class ExcelReader():
    
    def __init__(self, setupDict, xlsFile=None):
        """
        Required setup parameters:
            filePath    = './data/'
            fileName    = "test.xls"
            sheetName   = "Sheet0"
            timeIdxList  = ("B1", "C1")
            spaceIdxList = ('A2', 'A10')
        
        """
        self.setup(**setupDict)
        if xlsFile is not None:
            self.xlsFile = xlsFile
        
    
#    def setSetup(self, **kwargs):
#        for key in kwargs.keys():
#            self.setup[key] = kwargs[key]

    def setup(self, **kwargs):
        for key in kwargs.keys():
            
            
            if 'Idx' in key:
                # conversion of index
                setattr(self, key, [excelIdx2PandasIdx(x) for x in kwargs[key]])
            else:
                setattr(self, key, kwargs[key])
        if hasattr(self,'df'):
            del self.df
#           
#        self.validate()
    
    def getAllSheetNames(self):
        xlFile = pd.ExcelFile(os.path.join(self.filePath, self.fileName))
        sheetNameList = xlFile.sheet_names
        xlFile.close()
        return sheetNameList

    def validate(self,):
        for attr in REQUIRED_SETUP_FIELDS:
            if not hasattr(self, attr):
                print(attr + " is missing")

    def gatherValue(self, excelIdx):
        if not hasattr(self, 'df'):
            if hasattr(self, 'xlsFile'):
                self.df = pd.read_excel(self.xlsFile, sheet_name=self.sheetName, header=None)
            else:
                self.df = pd.read_excel(os.path.join(self.filePath, self.fileName), sheet_name=self.sheetName, header=None)
        elif core.config.DEBUG:
            print('use loaded df')
        return self.df.iloc[excelIdx2PandasIdx(excelIdx)]             

    def getAllData(self):
        return self.df
        
    def openFile(self):
        import os
        os.system('libreoffice ' + os.path.join(self.filePath, self.fileName)) 
                        
    def gatherData(self, load=True):

        if load:
            self.df = pd.read_excel(os.path.join(self.filePath, self.fileName), sheet_name=self.sheetName, header=None)
#            print('df loaded')
        elif core.config.DEBUG:
            print('use loaded df')

#        print(self.spaceIdxList)
#        print(self.timeIdxList)
        
        if len(self.timeIdxList) >1:
            timeIdx = self.df.iloc[self.timeIdxList[0][0],self.timeIdxList[0][1]:self.timeIdxList[1][1]+1]
            columns = self.df.columns[self.timeIdxList[0][1]:self.timeIdxList[1][1]+1]
#            print(timeIdx)
        else:
            timeIdx = self.df.iloc[self.timeIdxList[0][0],[self.timeIdxList[0][1]]]
            columns = self.df.columns[[self.timeIdxList[0][1]]]

        if len(self.spaceIdxList) >1:
            regions   = self.df.iloc[self.spaceIdxList[0][0]:self.spaceIdxList[1][0]+1,self.spaceIdxList[0][1]]
            index = self.df.index[self.spaceIdxList[0][0]:self.spaceIdxList[1][0]+1]
        else:
            regions   = self.df.iloc[[self.spaceIdxList[0][0]],self.spaceIdxList[0][1]]
            index = self.df.index[[self.spaceIdxList[0][0]]]
            
        if len(index) == 1:
#            print(index)
#            print(columns)
#            print(regions)
#            print(timeIdx)

            if len(columns) == 1:
                # extraction of a single value    
                data    = self.df.loc[index,columns]
                return pd.DataFrame(data.values, columns=regions, index=timeIdx).T
            
            else:
                # extraction of time serie
                data    = self.df.iloc[self.spaceIdxList[0][0],self.timeIdxList[0][1]:self.timeIdxList[1][1]+1]
                #print(type(data))
                #print(data.index)
                return pd.DataFrame(data.values, columns = regions, index = timeIdx).T
#                return pd.Series(data.values, index=columns) 
            
        else:
        
            if len(columns) == 1:
                # extraction of spatial series
                data    = self.df.loc[index,columns]
                
                return pd.DataFrame(data.values, columns=regions, index=timeIdx).T
#                return pd.Series(data.values, index=index) 
            else:
                # extraction of both spatial and time data
#                print(self.df)
#                print(index)
#                print(timeIdx)
#                print(columns)
                data    = self.df.loc[index,columns]
#                print(data)
                
                return pd.DataFrame(data.values, columns=timeIdx, index=regions)


  
class ExcelWriter():
    """
    More complex excel interface
    Author: Andreas Geiges
    """
    def __init__(self, setup=None, 
                 fileName=None, 
                 overwrite=False, 
                 setupSheet = 'INPUT',
                 interpolate=False):
        self.overwrite = overwrite # overwriting values
        self.setupSheet= setupSheet
        if setup:
            self.setup = setup
            fileSetup = True
        else:
            self.setup = dict()
            self.setup['fileName'] = fileName
            fileSetup = False
    
        self.interpolate = interpolate
    
    def close(self):
        self.wb.close()
        
        
    def getSetups(self):
        self.wb_read = load_workbook(self.setup['fileName'], data_only=True)
        self.wb = load_workbook(self.setup['fileName'])
        setup = dict()
        setup['fileName'] = self.setup['fileName']
        
        # new using pandas
        mapping = pd.read_excel(self.setup['fileName'], sheet_name=self.setupSheet)
        mapping.columns = [x.lower() for x in mapping.columns]
        
        
        setupDict = {'sheetName'     : 'sheetname',
                     'timeIdxList'   : 'time',
                     'spaceIdxList'  : 'region',
                     'dataID'        : 'variable',
                     'unit'          : 'unit',
#                     'unitTo'        : 'unitto'
                     }
        
        for i,setupMapp in mapping.iterrows():
            print(setupMapp)
            for key in setupDict.keys():
                setup[key] = str(setupMapp[setupDict[key]])
                if setup[key] == 'nan':
                    setup[key] = None
            yield(dict(setup))
        # old 
#        setupSheet = self.wb[self.setupSheet]
#        for row in setupSheet.rows:
#            if row[0].value == 'sheet':
#                continue
#            
#            setup['sheetName']   = row[0].value
#            setup['timeIdxList']  = row[1].value
#            setup['spaceIdxList'] = row[2].value
#            setup['dataID']      = row[3].value
#            setup['unit']        = row[4].value
#            
#            for key in setup.keys():
#                if isinstance(setup[key],str):
#                    setup[key] = setup[key].replace('\ufeff','')
#            yield(setup)
        
    def insert_data(self):
        replaceDict = {'EU': 'EU28'}
        from shutil import copyfile
        from copy import copy
        
        # Copy old file with "_filled" extension and load it
        if '.xlsx' in self.setup['fileName']:
            saveFileName = self.setup['fileName'].replace('.xlsx','_filled.xlsx')
        elif '.xls' in self.setup['fileName']:
            saveFileName = self.setup['fileName'].replace('.xls','_filled.xlsx')
        else:
            print('file extention not recognized')
        copyfile(self.setup['fileName'], saveFileName)
        self.setup['fileName'] = saveFileName
#        wb = load_workbook(self.setup['fileName'])
        
        
        # pre-load all valid spatial IDS
        validSpatialIDs = mapp.getValidSpatialIDs()
        
        # create list of setups
        self.setupList = list()
        
        #loop over all defined inputs in the INPUT sheet mapping (rows)
        for setup in self.getSetups():
            
            #ensure input is at all times a string
            if isinstance(setup['timeIdxList'], int):
                setup['timeIdxList'] = str(setup['timeIdxList'])
            if isinstance(setup['spaceIdxList'], int):
                setup['spaceIdxList'] = str(setup['spaceIdxList'])
            
            if config.DEBUG:
                print(setup)
            
            #add a copy of setup to list
            self.setupList.append(copy(setup))
            
            args =  None, None, None, None, None
            wksSheet_read  = self.wb_read[setup['sheetName']]
            wksSheet_write = self.wb[setup['sheetName']]
            # loop overall setups and collect dataIDs for pre-loading
            tableIDs = list()
            for args in _iterData(setup['dataID'], wksSheet_read, *args):
                
                if config.DEBUG:
                    print(args)
                if (~pd.isna(args[4])) and (args[4] is not None):
                    tableIDs.append(args[4])
            
            if config.DEBUG:
                print(tableIDs)
                
            #load all tables from database
            tables = core.DB.getTables(tableIDs)
            
            if setup['unit'] is not None:
                 for tableKey in tables.keys():
                     table = tables[tableKey]
                     tables[tableKey]= table.convert(setup['unit'])
                     
            if self.interpolate:
                 for tableKey in tables.keys():
                     for col in list(range(tables[tableKey].columns.min(),tables[tableKey].columns.max()+1)):
                         if col not in tables[tableKey].columns:
                             tables[tableKey].loc[:,col] = np.nan
                     tables[tableKey] = tables[tableKey].interpolate()    
                     
#            wksSheet = load_workbook(setup['fileName'], data_only=True)[setup['sheetName']]
            
            args =  [None, None, None, None, None]
            iCount = 0
            
            #iterate over time index list
            for argsTime in _iterTime(setup['timeIdxList'], wksSheet_read, *args):
                if argsTime[TI_ARG] is None:
                    print('No time defintion found')
                    continue
                
                #iterate over space index list
                for argsSpace in _iterSpace(setup['spaceIdxList'], wksSheet_read, *argsTime):
                    if argsSpace[SP_ARG] is None:
                        print('not spacial defintion found')
                        continue
                    
                    
                    if argsSpace[SP_ARG] not in validSpatialIDs:
#                        print('not time defintion found')
#                        print(argsSpace[SP_ARG])
                        if argsSpace[SP_ARG] in replaceDict.keys():
                            newID = replaceDict[argsSpace[SP_ARG]]
                        else:
                            newID = util.getCountryISO(argsSpace[SP_ARG])
                        

                        if newID in validSpatialIDs:
                            argsSpace[SP_ARG] = newID
                            print(argsSpace[SP_ARG])
                        else:
                            print(argsSpace[SP_ARG] + ' not found')
                            continue
                        
                    #print(argsSpace[SP_ARG])
                    #iterate over all data indices
                    for argData in _iterData(setup['dataID'], wksSheet_read, *argsSpace):
                        if argData[DT_ARG] is None:
                            print('not data defintion found')
                            continue

                        try:
                            value = tables[argData[DT_ARG]].loc[argData[SP_ARG],int(argData[TI_ARG])]
                            self._writeValueNew(wksSheet_write, xlsRow=argData[0], xlsCol=argData[1], value=value)
                            
#                            print('success')
                            iCount +=1
                            
                        except Exception as e:
                            if config.DEBUG:
                                print('failed with agruments:' + str(argData))
                                print(e)
                                #import pdb
                                #pdb.set_trace()
                            pass

            self.wb.save(self.setup['fileName'])
            print(setup)
            print('{} items inserted'.format(iCount))
            
            self.wb.close()

    def _writeValueNew(self, wrkSheet, xlsRow, xlsCol, value):
        if self.overwrite or pd.isna(wrkSheet.cell(row=xlsRow, column=xlsCol).value):
            print(value)
            wrkSheet.cell(row=xlsRow, column=xlsCol, value = value)
        
#    def _writeMultipleIndicators(self, setup):
#        ws_readValues = load_workbook(self.setup['fileName'], data_only=True)[setup['sheetName']]
#        try:
#            tableIDs = [cell.value for cell in ws_readValues[setup['dataID']]]
#        except:
#            tableIDs = [cell[0].value for cell in ws_readValues[setup['dataID']]]
#            
#            tables = dt.getTables(tableIDs)
#            
#            REG_FIND_ROWS.search(setup['timeIdxList'])
            
